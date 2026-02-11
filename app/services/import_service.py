import csv
import io
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Course, ImportLog, Student, Teacher


# 仅允许导入的表
ALLOWED_TABLES = {
    "student": Student,
    "teacher": Teacher,
    "course": Course,
}

# 导入时排除的系统字段
EXCLUDED_COLUMNS = {"id", "created_at", "updated_at", "created_by", "updated_by", "is_deleted"}


# 执行导入（校验失败不落库）
def import_data(
    table_name: str,
    filename: str,
    content: bytes,
    db: Session,
    admin_id: int,
) -> dict[str, Any]:
    def _helper_parse_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes"}
        return False

    def _helper_convert_value(column: Any, value: Any) -> Any:
        if isinstance(value, str) and value.strip() == "":
            return None

        try:
            python_type = column.type.python_type
        except (NotImplementedError, AttributeError):
            return value

        if python_type is bool:
            return _helper_parse_bool(value)
        if python_type is int:
            return int(value)
        if python_type is float:
            return float(value)
        if python_type is date:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                return date.fromisoformat(value)
        if python_type is datetime:
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                return datetime.fromisoformat(value)
        if python_type is str:
            return str(value)
        return python_type(value)

    if table_name not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table for import")

    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    ext = (filename or "").lower()
    if ext.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows: list[tuple[int, dict[str, Any]]] = []
        for idx, row in enumerate(reader, start=2):
            if not row:
                continue
            if all(value is None or str(value).strip() == "" for value in row.values()):
                continue
            rows.append((idx, row))
    elif ext.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        sheet_rows = list(sheet.iter_rows(values_only=True))
        if not sheet_rows:
            rows = []
        else:
            headers = [str(cell).strip() if cell is not None else "" for cell in sheet_rows[0]]
            rows = []
            for idx, row in enumerate(sheet_rows[1:], start=2):
                if row is None:
                    continue
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                row_dict: dict[str, Any] = {}
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row[col_idx] if col_idx < len(row) else None
                rows.append((idx, row_dict))
    else:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX is supported")

    model = ALLOWED_TABLES[table_name]
    allowed: dict[str, Any] = {}
    required: set[str] = set()
    for column in model.__table__.columns:
        if column.name in EXCLUDED_COLUMNS:
            continue
        allowed[column.name] = column
        is_autoincrement_pk = bool(column.primary_key and column.autoincrement is True)
        if (
            not column.nullable
            and column.default is None
            and column.server_default is None
            and not is_autoincrement_pk
        ):
            required.add(column.name)

    errors: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []
    for row_index, row in rows:
        row_errors: list[dict[str, Any]] = []
        data: dict[str, Any] = {}

        for field in required:
            raw = row.get(field)
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                row_errors.append({"row": row_index, "field": field, "message": "required"})

        for field, column in allowed.items():
            if field not in row:
                continue
            raw = row.get(field)
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue
            try:
                data[field] = _helper_convert_value(column, raw)
            except Exception:
                row_errors.append({"row": row_index, "field": field, "message": "invalid type"})

        if row_errors:
            errors.extend(row_errors)
            continue

        data["created_by"] = admin_id
        data["updated_by"] = admin_id
        data["is_deleted"] = False
        records.append(data)

    total_rows = len(rows)
    failed_rows = len({error["row"] for error in errors})
    success_rows = total_rows - failed_rows

    if errors:
        log = ImportLog(
            table_name=table_name,
            filename=filename,
            total_rows=total_rows,
            success_rows=0,
            failed_rows=failed_rows,
            status="failed",
            error_summary=str(errors[:10]),
            created_by=admin_id,
        )
        db.add(log)
        db.commit()
        return {
            "summary": {
                "table": table_name,
                "filename": filename,
                "total": total_rows,
                "success": 0,
                "failed": failed_rows,
            },
            "errors": errors,
        }

    try:
        items = [model(**record) for record in records]
        db.add_all(items)
        log = ImportLog(
            table_name=table_name,
            filename=filename,
            total_rows=total_rows,
            success_rows=success_rows,
            failed_rows=0,
            status="success",
            error_summary=None,
            created_by=admin_id,
        )
        db.add(log)
        db.commit()
    except Exception as exc:
        db.rollback()
        log = ImportLog(
            table_name=table_name,
            filename=filename,
            total_rows=total_rows,
            success_rows=0,
            failed_rows=total_rows,
            status="failed",
            error_summary=str(exc),
            created_by=admin_id,
        )
        db.add(log)
        db.commit()
        raise HTTPException(status_code=500, detail="Import failed")

    return {
        "summary": {
            "table": table_name,
            "filename": filename,
            "total": total_rows,
            "success": success_rows,
            "failed": 0,
        },
        "errors": [],
    }

